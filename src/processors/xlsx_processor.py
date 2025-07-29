"""
XLSX document processor
"""

import copy
from typing import Any, Dict, List

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .base import BaseDocumentProcessor


class XLSXProcessor(BaseDocumentProcessor):
    """Processor for XLSX documents"""

    def extract_text(self, file_path: str) -> Dict[str, Any]:
        """Extract text content from XLSX while preserving structure"""
        wb = load_workbook(file_path)
        content = {"worksheets": []}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_content = {"sheet_name": sheet_name, "cells": []}

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip():
                        # Only extract text/string values
                        if isinstance(cell.value, (str, int, float)):
                            cell_text = str(cell.value).strip()
                            if cell_text:
                                sheet_content["cells"].append(
                                    {
                                        "row": cell.row,
                                        "column": cell.column,
                                        "coordinate": cell.coordinate,
                                        "text": cell_text,
                                        "original_text": cell_text,
                                        "data_type": cell.data_type,
                                        "font": (
                                            {
                                                "name": cell.font.name,
                                                "size": cell.font.size,
                                                "bold": cell.font.bold,
                                                "italic": cell.font.italic,
                                            }
                                            if cell.font
                                            else None
                                        ),
                                    }
                                )

            content["worksheets"].append(sheet_content)

        wb.close()
        return content

    def reconstruct_document(
        self, original_path: str, translated_content: Dict[str, Any], output_path: str
    ) -> str:
        """Reconstruct XLSX by modifying original workbook"""
        wb = load_workbook(original_path)

        for sheet_data in translated_content["worksheets"]:
            sheet_name = sheet_data["sheet_name"]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Update cells with translated content
                for cell_data in sheet_data["cells"]:
                    coordinate = cell_data["coordinate"]
                    cell = ws[coordinate]
                    cell.value = cell_data["text"]

                    # Preserve formatting if available
                    if cell_data.get("font"):
                        font_info = cell_data["font"]
                        cell.font = Font(
                            name=font_info.get("name", "Calibri"),
                            size=font_info.get("size", 11),
                            bold=font_info.get("bold", False),
                            italic=font_info.get("italic", False),
                        )

        wb.save(output_path)
        wb.close()
        return output_path

    def get_translatable_texts(self, extracted_content: Dict[str, Any]) -> List[str]:
        """Extract all translatable texts from XLSX content"""
        texts = []

        for sheet_data in extracted_content["worksheets"]:
            for cell_data in sheet_data["cells"]:
                # Only translate text cells, not numbers or formulas
                if isinstance(cell_data["text"], str) and cell_data["text"].strip():
                    # Simple check to avoid translating numbers
                    try:
                        float(cell_data["text"])
                        # Skip if it's a number
                        continue
                    except ValueError:
                        # It's text, add to translation list
                        texts.append(cell_data["text"])

        return texts

    def apply_translations(
        self, extracted_content: Dict[str, Any], translations: List[str]
    ) -> Dict[str, Any]:
        """Apply translations to XLSX content structure"""
        translated_content = copy.deepcopy(extracted_content)
        translation_idx = 0

        for sheet_data in translated_content["worksheets"]:
            for cell_data in sheet_data["cells"]:
                # Only apply translations to text cells
                if isinstance(cell_data["text"], str) and cell_data["text"].strip():
                    try:
                        float(cell_data["text"])
                        # Skip numbers
                        continue
                    except ValueError:
                        # It's text, apply translation
                        if translation_idx < len(translations):
                            cell_data["text"] = translations[translation_idx]
                            translation_idx += 1

        return translated_content
